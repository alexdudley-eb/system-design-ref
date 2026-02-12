import sys
import os
from openpyxl import load_workbook
from sqlalchemy.orm import Session

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from database import SessionLocal, init_db
from models import Tool, ToolDeep

def clean_cell_value(cell_value):
    if cell_value is None:
        return None
    if isinstance(cell_value, str):
        return cell_value.strip() if cell_value.strip() else None
    return str(cell_value)

def get_header_flexible(headers, *variants):
    """Try multiple header name variants, return the column index if found."""
    for variant in variants:
        if variant in headers:
            return headers[variant]
    return None

def import_quick_reference(ws, db: Session):
    print("Importing Quick Reference sheet...")
    
    header_row = 2
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            headers[cell.value.strip()] = col_idx
    
    print(f"Found headers: {list(headers.keys())}")
    
    tools_added = 0
    for row_idx in range(3, ws.max_row + 1):
        row = ws[row_idx]
        
        name = clean_cell_value(row[headers.get("Technology", 1) - 1].value)
        if not name:
            continue
        
        category = clean_cell_value(row[headers.get("Category", 2) - 1].value)
        if not category:
            continue
        
        existing_tool = db.query(Tool).filter(Tool.name == name).first()
        if existing_tool:
            print(f"Skipping duplicate: {name}")
            continue
        
        tool = Tool(
            name=name,
            category=category,
            cap_leaning=clean_cell_value(row[headers.get("CAP Leaning", 3) - 1].value),
            consistency_model=clean_cell_value(row[headers.get("Consistency Model", 4) - 1].value),
            interview_oneliner=clean_cell_value(row[headers.get("Interview One-liner", 5) - 1].value),
            best_for=clean_cell_value(row[headers.get("Best Used For", 6) - 1].value),
            avoid_when=clean_cell_value(row[headers.get("Avoid When", 7) - 1].value),
            tradeoffs=clean_cell_value(row[headers.get("Key Tradeoffs", 8) - 1].value),
            scaling_pattern=clean_cell_value(row[headers.get("Scaling Pattern", 9) - 1].value),
            official_docs_url=clean_cell_value(row[(get_header_flexible(headers, "Official Docs", "Docs_URL", "official_docs_url") or 10) - 1].value),
            deep_dive_url_1=clean_cell_value(row[(get_header_flexible(headers, "Deep Dive 1", "Deep_Dive_1", "deep_dive_url_1") or 11) - 1].value),
            deep_dive_url_2=clean_cell_value(row[(get_header_flexible(headers, "Deep Dive 2", "Deep_Dive_2", "deep_dive_url_2") or 12) - 1].value),
            aws_only=1 if "AWS" in category or "Amazon" in name else 0
        )
        
        db.add(tool)
        tools_added += 1
        print(f"  Added: {name}")
    
    db.commit()
    print(f"Imported {tools_added} tools from Quick Reference")

def find_tool_by_name(db: Session, name: str):
    tool = db.query(Tool).filter(Tool.name == name).first()
    if tool:
        return tool
    
    all_tools = db.query(Tool).all()
    
    for t in all_tools:
        if name.lower() in t.name.lower() or t.name.lower() in name.lower():
            return t
    
    return None

def import_deep_study(ws, db: Session):
    print("Importing Deep Study sheet...")
    
    all_tools = db.query(Tool).all()
    print(f"Available tools in database: {len(all_tools)} tools")
    
    header_row = 2
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            headers[cell.value.strip()] = col_idx
    
    print(f"Found headers: {list(headers.keys())}")
    
    deep_studies_added = 0
    for row_idx in range(3, ws.max_row + 1):
        row = ws[row_idx]
        
        name = clean_cell_value(row[headers.get("Technology", 1) - 1].value)
        if not name:
            continue
        
        tool = find_tool_by_name(db, name)
        if not tool:
            print(f"  ❌ Tool not found for deep study: '{name}'")
            continue
        
        if tool.name != name:
            print(f"  🔗 Matched '{name}' -> '{tool.name}'")
        
        existing_deep = db.query(ToolDeep).filter(ToolDeep.tool_id == tool.id).first()
        if existing_deep:
            print(f"Skipping duplicate deep study: {name}")
            continue
        
        tool_deep = ToolDeep(
            tool_id=tool.id,
            failure_modes=clean_cell_value(row[headers.get("Failure Modes", 2) - 1].value),
            multi_region_notes=clean_cell_value(row[headers.get("Multi-Region / DR", 3) - 1].value),
            tuning_gotchas=clean_cell_value(row[headers.get("Tuning Gotchas", 4) - 1].value),
            observability_signals=clean_cell_value(row[headers.get("Observability", 5) - 1].value),
            alternatives=clean_cell_value(row[headers.get("Alternatives", 6) - 1].value),
            interview_prompts=clean_cell_value(row[headers.get("Interview Prompts", 7) - 1].value)
        )
        
        db.add(tool_deep)
        deep_studies_added += 1
        print(f"  ✅ Added deep study for: {name}")
    
    db.commit()
    print(f"Imported {deep_studies_added} deep studies")

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Import system design data from XLSX")
    parser.add_argument("--refresh", action="store_true", help="Clear existing data before import")
    parser.add_argument("--file", type=str, default="../data/system-design-tech-cheatsheet.xlsx", 
                       help="Path to XLSX file")
    args = parser.parse_args()
    
    init_db()
    
    db = SessionLocal()
    
    if args.refresh:
        print("Clearing existing data...")
        db.query(ToolDeep).delete()
        db.query(Tool).delete()
        db.commit()
    
    xlsx_path = os.path.join(os.path.dirname(__file__), args.file)
    
    if not os.path.exists(xlsx_path):
        print(f"Error: XLSX file not found at {xlsx_path}")
        print("Please place the system-design-tech-cheatsheet.xlsx file in the data/ directory")
        sys.exit(1)
    
    print(f"Loading workbook from {xlsx_path}...")
    wb = load_workbook(xlsx_path)
    
    if "Quick Reference" in wb.sheetnames:
        ws_quick = wb["Quick Reference"]
        import_quick_reference(ws_quick, db)
    else:
        print("Warning: 'Quick Reference' sheet not found")
    
    if "Deep Study" in wb.sheetnames:
        ws_deep = wb["Deep Study"]
        import_deep_study(ws_deep, db)
    else:
        print("Warning: 'Deep Study' sheet not found")
    
    db.close()
    print("\nImport completed successfully!")
    print(f"Database location: {os.path.join(os.path.dirname(__file__), 'system_design_ref.db')}")

if __name__ == "__main__":
    main()
